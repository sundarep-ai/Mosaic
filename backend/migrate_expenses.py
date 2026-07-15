"""
Migrate expenses from an existing .xlsx or .csv file into the Mosaic SQLite database.

Usage:
    python migrate_expenses.py path/to/your/expenses.xlsx --date-format DD/MM/YYYY
    python migrate_expenses.py path/to/your/expenses.csv --date-format YYYY-MM-DD

Expected columns (header matching is case-insensitive and flexible):
    Date, Description, Amount, Category, Paid By, Split Method

--date-format is required: a string date like "01/02/2026" is ambiguous
(Jan 2 or Feb 1) and guessing silently produces day/month-swapped dates for
some rows and not others, leaving self-inconsistent history with no error.
Dates already stored as real Excel date cells (not text) are read directly —
the flag only matters for string date cells (always the case for .csv).

The whole file is validated before anything is written: if any row fails
validation, nothing is imported. A partial import (some rows in, some
silently dropped mid-run) is worse than refusing to run at all.
"""

import argparse
import csv
import sys
from datetime import datetime, date as date_cls
from decimal import InvalidOperation
from pathlib import Path

import openpyxl
from pydantic import ValidationError
from sqlalchemy import func
from sqlmodel import Session, select

from database import engine, create_db_and_tables
from models import Expense
from routes.expenses import VALID_CATEGORIES, _VALID_CATEGORIES_CASEFOLD
from users import get_all_users

# Maps the app's own date-format vocabulary to strptime patterns. Kept
# deliberately small and explicit (no guessing) — see module docstring.
DATE_FORMAT_MAP = {
    "YYYY-MM-DD": "%Y-%m-%d",
    "MM/DD/YYYY": "%m/%d/%Y",
    "DD/MM/YYYY": "%d/%m/%Y",
}


def normalize(header: str) -> str:
    return header.strip().lower().replace(" ", "_").replace("-", "_")


def _load_rows(filepath: str) -> tuple[list[str], list[tuple]]:
    """Return (normalized_headers, list_of_value_tuples) for .xlsx or .csv."""
    suffix = Path(filepath).suffix.lower()
    if suffix == ".csv":
        with open(filepath, newline="", encoding="utf-8-sig") as fh:
            reader = csv.reader(fh)
            raw_headers = next(reader)
            headers = [normalize(h) for h in raw_headers]
            rows = [tuple(row) for row in reader]
        return headers, rows
    else:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [normalize(str(cell.value)) for cell in ws[1]]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        return headers, rows


def _parse_date(raw, strptime_fmt: str, row_num: int, errors: list[str]):
    """Return a date for a cell that's either a native Excel date or a string
    in the explicitly-requested format. Appends to `errors` and returns None
    on failure rather than raising, so the caller can collect every bad row
    before deciding whether to abort.
    """
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date_cls):
        return raw
    try:
        return datetime.strptime(str(raw).strip(), strptime_fmt).date()
    except ValueError:
        errors.append(f"Row {row_num}: unparseable date {raw!r} for format {strptime_fmt!r}")
        return None


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Migrate expenses into Mosaic's database.")
    parser.add_argument("filepath", help="Path to the .xlsx or .csv file")
    parser.add_argument(
        "--date-format",
        required=True,
        choices=sorted(DATE_FORMAT_MAP),
        help="Format of string date cells in the sheet (ignored for native Excel date cells).",
    )
    return parser


def migrate(filepath: str, date_format: str) -> None:
    create_db_and_tables()

    suffix = Path(filepath).suffix.lower()
    if suffix not in (".xlsx", ".csv"):
        print(f"ERROR: Unsupported file type '{suffix}'. Use .xlsx or .csv.")
        sys.exit(1)

    strptime_fmt = DATE_FORMAT_MAP[date_format]

    headers, rows = _load_rows(filepath)

    # Map header keywords to field names
    col_map: dict[str, int] = {}
    for i, h in enumerate(headers):
        if "date" in h:
            col_map.setdefault("date", i)
        elif "desc" in h:
            col_map.setdefault("description", i)
        elif "amount" in h:
            col_map.setdefault("amount", i)
        elif "categ" in h:
            col_map.setdefault("category", i)
        elif "paid" in h:
            col_map.setdefault("paid_by", i)
        elif "split" in h:
            col_map.setdefault("split_method", i)

    required = {"date", "description", "amount", "category", "paid_by", "split_method"}
    missing = required - col_map.keys()
    if missing:
        print(f"ERROR: Could not map columns: {missing}")
        print(f"  Found headers: {headers}")
        sys.exit(1)

    with Session(engine) as session:
        # Validate that every paid_by value in the sheet matches a registered display name
        registered_display_names = {u.display_name for u in get_all_users(session)}
        if not registered_display_names:
            print("ERROR: No registered users found in the database. Please set up your account(s) before migrating.")
            sys.exit(1)

        sheet_paid_by = {
            str(row[col_map["paid_by"]]).strip()
            for row in rows
            if row and row[col_map["paid_by"]] not in (None, "")
        }
        unknown = sheet_paid_by - registered_display_names
        if unknown:
            print("ERROR: The following 'Paid By' values in the sheet do not match any registered user display name:")
            for name in sorted(unknown):
                print(f"  - {name!r}")
            print(f"Registered display names: {sorted(registered_display_names)}")
            print("Fix the sheet or register the missing users before migrating.")
            sys.exit(1)

        existing_count = session.exec(select(func.count(Expense.id))).one()
        if existing_count > 0:
            answer = input(
                f"WARNING: Database already has {existing_count} expenses. "
                "Continue and add new rows? [y/N]: "
            )
            if answer.strip().lower() != "y":
                print("Aborted.")
                sys.exit(0)

        # Every split method that could legitimately appear across this app's
        # history, regardless of the *current* app_mode toggle (which is a
        # global runtime setting, not a per-row historical marker).
        valid_split_methods = {"50/50", "Personal"} | {
            f"100% {name}" for name in registered_display_names
        }

        # ── Pass 1: validate every row before touching the database ──
        errors: list[str] = []
        validated: list[Expense] = []

        for row_num, row in enumerate(rows, start=2):
            if not row or row[col_map["date"]] in (None, ""):
                continue

            row_ok = True

            date_val = _parse_date(row[col_map["date"]], strptime_fmt, row_num, errors)
            if date_val is None:
                row_ok = False

            category_cell = row[col_map["category"]]
            category = "" if category_cell is None else str(category_cell).strip()
            if not category:
                errors.append(f"Row {row_num}: category cannot be empty")
                row_ok = False
            else:
                canonical = _VALID_CATEGORIES_CASEFOLD.get(category.casefold())
                if canonical and canonical != category:
                    errors.append(
                        f"Row {row_num}: category {category!r} differs only in "
                        f"case from existing category {canonical!r} — use {canonical!r}"
                    )
                    row_ok = False

            split_method = str(row[col_map["split_method"]]).strip()
            if split_method not in valid_split_methods:
                errors.append(f"Row {row_num}: unknown split_method {split_method!r}")
                row_ok = False

            if not row_ok:
                continue  # already recorded above; nothing more to build for this row

            paid_by = str(row[col_map["paid_by"]]).strip()

            try:
                expense = Expense.model_validate({
                    "date": date_val,
                    "description": str(row[col_map["description"]]).strip(),
                    "amount": row[col_map["amount"]],
                    "category": category,
                    "paid_by": paid_by,
                    "split_method": split_method,
                    "user_id": None,
                })
            except (ValidationError, InvalidOperation, ValueError, TypeError) as e:
                errors.append(f"Row {row_num}: {e}")
                continue

            if expense.amount == 0:
                errors.append(f"Row {row_num}: amount cannot be zero")
                continue

            validated.append(expense)

        if errors:
            print(f"ERROR: {len(errors)} row(s) failed validation. Nothing was imported.")
            for e in errors:
                print(f"  - {e}")
            sys.exit(1)

        # ── Pass 2: every row validated — commit as a single unit ──
        for expense in validated:
            session.add(expense)
        session.commit()
        print(f"Successfully migrated {len(validated)} expenses into mosaic.db")


if __name__ == "__main__":
    args = build_arg_parser().parse_args()
    migrate(args.filepath, args.date_format)
