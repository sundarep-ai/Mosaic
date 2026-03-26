"""
Migrate expenses from an existing .xlsx file into the TallyUs SQLite database.

Usage:
    python migrate_xlsx.py path/to/your/expenses.xlsx

Expected columns (header matching is case-insensitive and flexible):
    Date, Description, Amount, Category, Paid By, Split Method
"""

import sys
from datetime import datetime
from decimal import Decimal

import openpyxl
from sqlalchemy import func
from sqlmodel import Session, select

from database import engine, create_db_and_tables
from models import Expense


def normalize(header: str) -> str:
    return header.strip().lower().replace(" ", "_").replace("-", "_")


def migrate(filepath: str) -> None:
    create_db_and_tables()

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    headers = [normalize(str(cell.value)) for cell in ws[1]]

    # Map header keywords to field names
    col_map: dict[str, int] = {}
    for i, h in enumerate(headers):
        if "date" in h:
            col_map.setdefault("date", i)
        elif "desc" in h:
            col_map.setdefault("description", i)
        elif "amount" in h:
            col_map.setdefault("amount", i)
        elif "categ" in h:
            col_map.setdefault("category", i)
        elif "paid" in h:
            col_map.setdefault("paid_by", i)
        elif "split" in h:
            col_map.setdefault("split_method", i)

    required = {"date", "description", "amount", "category", "paid_by", "split_method"}
    missing = required - col_map.keys()
    if missing:
        print(f"ERROR: Could not map columns: {missing}")
        print(f"  Found headers: {headers}")
        sys.exit(1)

    with Session(engine) as session:
        existing_count = session.exec(select(func.count(Expense.id))).one()
        if existing_count > 0:
            answer = input(
                f"WARNING: Database already has {existing_count} expenses. "
                "Continue and add new rows? [y/N]: "
            )
            if answer.strip().lower() != "y":
                print("Aborted.")
                sys.exit(0)

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[col_map["date"]] is None:
                continue

            date_val = row[col_map["date"]]
            if isinstance(date_val, str):
                # Try common date formats
                for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y"):
                    try:
                        date_val = datetime.strptime(date_val, fmt).date()
                        break
                    except ValueError:
                        continue
                else:
                    print(f"WARNING: Skipping row with unparseable date: {date_val}")
                    continue
            elif isinstance(date_val, datetime):
                date_val = date_val.date()

            expense = Expense(
                date=date_val,
                description=str(row[col_map["description"]]).strip(),
                amount=Decimal(str(row[col_map["amount"]])),
                category=str(row[col_map["category"]]).strip(),
                paid_by=str(row[col_map["paid_by"]]).strip(),
                split_method=str(row[col_map["split_method"]]).strip(),
                user_id=None,
            )
            session.add(expense)
            count += 1

        session.commit()
        print(f"Successfully migrated {count} expenses into tallyus.db")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python migrate_xlsx.py <path_to_xlsx>")
        sys.exit(1)
    migrate(sys.argv[1])
