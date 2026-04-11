"""
Migrate income entries from an existing .xlsx file into the Mosaic SQLite database.

Usage:
    python migrate_income_xlsx.py path/to/your/income.xlsx

Expected columns (header matching is case-insensitive and flexible):
    Date, Amount, Source, Display Name
    Notes  (optional)

Valid sources: "Salary / Wages", "Freelance / Side Income", "Other"
"""

import sys
from datetime import datetime
from decimal import Decimal, InvalidOperation

import openpyxl
from sqlalchemy import func
from sqlmodel import Session, select

from database import engine, create_db_and_tables
from models import Income, VALID_INCOME_SOURCES
from users import get_all_users


def normalize(header: str) -> str:
    return header.strip().lower().replace(" ", "_").replace("-", "_")


def migrate(filepath: str) -> None:
    create_db_and_tables()

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    headers = [normalize(str(cell.value)) for cell in ws[1]]

    # Map header keywords to field names
    col_map: dict[str, int] = {}
    for i, h in enumerate(headers):
        if "date" in h:
            col_map.setdefault("date", i)
        elif "amount" in h:
            col_map.setdefault("amount", i)
        elif "source" in h:
            col_map.setdefault("source", i)
        elif "display" in h or "user" in h:
            col_map.setdefault("display_name", i)
        elif "note" in h:
            col_map.setdefault("notes", i)

    required = {"date", "amount", "source", "display_name"}
    missing = required - col_map.keys()
    if missing:
        print(f"ERROR: Could not map columns: {missing}")
        print(f"  Found headers: {headers}")
        sys.exit(1)

    valid_sources = sorted(VALID_INCOME_SOURCES)

    with Session(engine) as session:
        # Validate that every display name in the sheet matches a registered user
        all_users = get_all_users(session)
        if not all_users:
            print("ERROR: No registered users found in the database. Please set up your account(s) before migrating.")
            sys.exit(1)

        registered_display_names = {u.display_name for u in all_users}
        # Build a lookup from display_name → username for use during row insertion
        display_name_to_username = {u.display_name: u.username for u in all_users}

        sheet_display_names = {
            str(row[col_map["display_name"]]).strip()
            for row in ws.iter_rows(min_row=2, values_only=True)
            if row and row[col_map["display_name"]] is not None
        }
        unknown = sheet_display_names - registered_display_names
        if unknown:
            print("ERROR: The following 'Display Name' values in the sheet do not match any registered user:")
            for name in sorted(unknown):
                print(f"  - {name!r}")
            print(f"Registered display names: {sorted(registered_display_names)}")
            print("Fix the sheet or register the missing users before migrating.")
            sys.exit(1)

        existing_count = session.exec(select(func.count(Income.id))).one()
        if existing_count > 0:
            answer = input(
                f"WARNING: Database already has {existing_count} income entries. "
                "Continue and add new rows? [y/N]: "
            )
            if answer.strip().lower() != "y":
                print("Aborted.")
                sys.exit(0)

        count = 0
        skipped = 0
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or row[col_map["date"]] is None:
                continue

            # --- Date ---
            date_val = row[col_map["date"]]
            if isinstance(date_val, str):
                for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y"):
                    try:
                        date_val = datetime.strptime(date_val, fmt).date()
                        break
                    except ValueError:
                        continue
                else:
                    print(f"WARNING: Row {row_num} — skipping unparseable date: {date_val!r}")
                    skipped += 1
                    continue
            elif isinstance(date_val, datetime):
                date_val = date_val.date()

            # --- Amount ---
            raw_amount = row[col_map["amount"]]
            try:
                amount_val = Decimal(str(raw_amount)).quantize(Decimal("0.01"))
                if amount_val <= 0:
                    raise ValueError("non-positive")
            except (InvalidOperation, ValueError):
                print(f"WARNING: Row {row_num} — skipping invalid amount: {raw_amount!r}")
                skipped += 1
                continue

            # --- Source ---
            source_val = str(row[col_map["source"]]).strip()
            if source_val not in VALID_INCOME_SOURCES:
                print(
                    f"WARNING: Row {row_num} — skipping unknown source: {source_val!r}. "
                    f"Valid values: {valid_sources}"
                )
                skipped += 1
                continue

            # --- Display Name → resolve to username ---
            raw_display = row[col_map["display_name"]]
            if raw_display is None:
                print(f"WARNING: Row {row_num} — skipping row with empty display name")
                skipped += 1
                continue
            display_name_val = str(raw_display).strip()
            if not display_name_val:
                print(f"WARNING: Row {row_num} — skipping row with empty display name")
                skipped += 1
                continue
            user_id_val = display_name_to_username[display_name_val]

            # --- Notes (optional) ---
            notes_val = None
            if "notes" in col_map and row[col_map["notes"]] is not None:
                notes_val = str(row[col_map["notes"]]).strip() or None

            income = Income(
                date=date_val,
                amount=amount_val,
                source=source_val,
                notes=notes_val,
                user_id=user_id_val,
            )
            session.add(income)
            count += 1

        session.commit()
        print(f"Successfully migrated {count} income entries into mosaic.db")
        if skipped:
            print(f"Skipped {skipped} rows due to validation errors (see warnings above).")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python migrate_income_xlsx.py <path_to_xlsx>")
        sys.exit(1)
    migrate(sys.argv[1])
