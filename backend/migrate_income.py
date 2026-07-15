"""
Migrate income entries from an existing .xlsx or .csv file into the Mosaic SQLite database.

Usage:
    python migrate_income.py path/to/your/income.xlsx --date-format DD/MM/YYYY
    python migrate_income.py path/to/your/income.csv --date-format YYYY-MM-DD

Expected columns (header matching is case-insensitive and flexible):
    Date, Amount, Source, Display Name
    Notes  (optional)

Valid sources: "Salary / Wages", "Freelance / Side Income", "Other"

--date-format is required: a string date like "01/02/2026" is ambiguous
(Jan 2 or Feb 1) and guessing silently produces day/month-swapped dates for
some rows and not others, leaving self-inconsistent history with no error.
Dates already stored as real Excel date cells (not text) are read directly —
the flag only matters for string date cells (always the case for .csv).

The whole file is validated before anything is written: if any row fails
validation, nothing is imported. A partial import (some rows in, some
silently dropped mid-run) is worse than refusing to run at all.
"""

import argparse
import csv
import sys
from datetime import datetime, date as date_cls
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from pydantic import ValidationError
from sqlalchemy import func
from sqlmodel import Session, select

from database import engine, create_db_and_tables
from models import Income, VALID_INCOME_SOURCES
from users import get_all_users

# Maps the app's own date-format vocabulary to strptime patterns. Kept
# deliberately small and explicit (no guessing) — see module docstring.
DATE_FORMAT_MAP = {
    "YYYY-MM-DD": "%Y-%m-%d",
    "MM/DD/YYYY": "%m/%d/%Y",
    "DD/MM/YYYY": "%d/%m/%Y",
}


def normalize(header: str) -> str:
    return header.strip().lower().replace(" ", "_").replace("-", "_")


def _load_rows(filepath: str) -> tuple[list[str], list[tuple]]:
    """Return (normalized_headers, list_of_value_tuples) for .xlsx or .csv."""
    suffix = Path(filepath).suffix.lower()
    if suffix == ".csv":
        with open(filepath, newline="", encoding="utf-8-sig") as fh:
            reader = csv.reader(fh)
            raw_headers = next(reader)
            headers = [normalize(h) for h in raw_headers]
            rows = [tuple(row) for row in reader]
        return headers, rows
    else:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [normalize(str(cell.value)) for cell in ws[1]]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        return headers, rows


def _parse_date(raw, strptime_fmt: str, row_num: int, errors: list[str]):
    """Return a date for a cell that's either a native Excel date or a string
    in the explicitly-requested format. Appends to `errors` and returns None
    on failure rather than raising, so the caller can collect every bad row
    before deciding whether to abort.
    """
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date_cls):
        return raw
    try:
        return datetime.strptime(str(raw).strip(), strptime_fmt).date()
    except ValueError:
        errors.append(f"Row {row_num}: unparseable date {raw!r} for format {strptime_fmt!r}")
        return None


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Migrate income entries into Mosaic's database.")
    parser.add_argument("filepath", help="Path to the .xlsx or .csv file")
    parser.add_argument(
        "--date-format",
        required=True,
        choices=sorted(DATE_FORMAT_MAP),
        help="Format of string date cells in the sheet (ignored for native Excel date cells).",
    )
    return parser


def migrate(filepath: str, date_format: str) -> None:
    create_db_and_tables()

    suffix = Path(filepath).suffix.lower()
    if suffix not in (".xlsx", ".csv"):
        print(f"ERROR: Unsupported file type '{suffix}'. Use .xlsx or .csv.")
        sys.exit(1)

    strptime_fmt = DATE_FORMAT_MAP[date_format]

    headers, rows = _load_rows(filepath)

    # Map header keywords to field names
    col_map: dict[str, int] = {}
    for i, h in enumerate(headers):
        if "date" in h:
            col_map.setdefault("date", i)
        elif "amount" in h:
            col_map.setdefault("amount", i)
        elif "source" in h:
            col_map.setdefault("source", i)
        elif "display" in h or "user" in h:
            col_map.setdefault("display_name", i)
        elif "note" in h:
            col_map.setdefault("notes", i)

    required = {"date", "amount", "source", "display_name"}
    missing = required - col_map.keys()
    if missing:
        print(f"ERROR: Could not map columns: {missing}")
        print(f"  Found headers: {headers}")
        sys.exit(1)

    valid_sources = sorted(VALID_INCOME_SOURCES)

    with Session(engine) as session:
        # Validate that every display name in the sheet matches a registered user
        all_users = get_all_users(session)
        if not all_users:
            print("ERROR: No registered users found in the database. Please set up your account(s) before migrating.")
            sys.exit(1)

        registered_display_names = {u.display_name for u in all_users}
        # Build a lookup from display_name → username for use during row insertion
        display_name_to_username = {u.display_name: u.username for u in all_users}

        sheet_display_names = {
            str(row[col_map["display_name"]]).strip()
            for row in rows
            if row and row[col_map["display_name"]] not in (None, "")
        }
        unknown = sheet_display_names - registered_display_names
        if unknown:
            print("ERROR: The following 'Display Name' values in the sheet do not match any registered user:")
            for name in sorted(unknown):
                print(f"  - {name!r}")
            print(f"Registered display names: {sorted(registered_display_names)}")
            print("Fix the sheet or register the missing users before migrating.")
            sys.exit(1)

        existing_count = session.exec(select(func.count(Income.id))).one()
        if existing_count > 0:
            answer = input(
                f"WARNING: Database already has {existing_count} income entries. "
                "Continue and add new rows? [y/N]: "
            )
            if answer.strip().lower() != "y":
                print("Aborted.")
                sys.exit(0)

        # ── Pass 1: validate every row before touching the database ──
        errors: list[str] = []
        validated: list[Income] = []

        for row_num, row in enumerate(rows, start=2):
            if not row or row[col_map["date"]] in (None, ""):
                continue

            row_ok = True

            date_val = _parse_date(row[col_map["date"]], strptime_fmt, row_num, errors)
            if date_val is None:
                row_ok = False

            # --- Amount --- validated up front (not via model_validate) so a
            # non-numeric cell reports a friendly row-numbered error instead of
            # an uncaught decimal.InvalidOperation (pydantic only catches
            # ValueError/TypeError/AssertionError raised inside a validator).
            raw_amount = row[col_map["amount"]]
            amount_val = None
            try:
                amount_val = Decimal(str(raw_amount)).quantize(Decimal("0.01"))
                if amount_val <= 0:
                    raise ValueError("non-positive")
            except (InvalidOperation, ValueError):
                errors.append(f"Row {row_num}: invalid amount {raw_amount!r}")
                row_ok = False

            source_val = str(row[col_map["source"]]).strip()
            if source_val not in VALID_INCOME_SOURCES:
                errors.append(
                    f"Row {row_num}: unknown source {source_val!r}. Valid values: {valid_sources}"
                )
                row_ok = False

            raw_display = row[col_map["display_name"]]
            display_name_val = str(raw_display).strip() if raw_display is not None else ""
            if not display_name_val:
                errors.append(f"Row {row_num}: empty display name")
                row_ok = False

            if not row_ok:
                continue  # already recorded above

            notes_val = None
            if "notes" in col_map and row[col_map["notes"]] is not None:
                notes_val = str(row[col_map["notes"]]).strip() or None

            try:
                income = Income.model_validate({
                    "date": date_val,
                    "amount": amount_val,
                    "source": source_val,
                    "notes": notes_val,
                    "user_id": display_name_to_username[display_name_val],
                })
            except ValidationError as e:
                errors.append(f"Row {row_num}: {e}")
                continue

            validated.append(income)

        if errors:
            print(f"ERROR: {len(errors)} row(s) failed validation. Nothing was imported.")
            for e in errors:
                print(f"  - {e}")
            sys.exit(1)

        # ── Pass 2: every row validated — commit as a single unit ──
        for income in validated:
            session.add(income)
        session.commit()
        print(f"Successfully migrated {len(validated)} income entries into mosaic.db")


if __name__ == "__main__":
    args = build_arg_parser().parse_args()
    migrate(args.filepath, args.date_format)
