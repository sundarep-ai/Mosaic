"""Tests for XLSX export endpoint."""

from datetime import date, timedelta
from io import BytesIO

import openpyxl

from conftest import make_expense, make_income, set_mode


def test_export_returns_xlsx(auth_client_a):
    auth_client_a.post("/api/expenses", json=make_expense())
    resp = auth_client_a.get("/api/export")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert "attachment" in resp.headers["content-disposition"]


def test_export_headers(auth_client_a):
    auth_client_a.post("/api/expenses", json=make_expense())
    resp = auth_client_a.get("/api/export")
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert headers == ["ID", "Date", "Description", "Amount", "Category", "Paid By", "Split Method"]


def test_export_contains_data(auth_client_a):
    auth_client_a.post("/api/expenses", json=make_expense(
        description="Walmart run", amount=55.50,
    ))
    resp = auth_client_a.get("/api/export")
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    assert ws.max_row == 2  # header + 1 data row
    assert ws.cell(2, 3).value == "Walmart run"
    assert ws.cell(2, 4).value == 55.5


def test_export_with_category_filter(auth_client_a):
    """Row count alone can't distinguish "kept the right row" from "kept the
    wrong one" — assert which row survived the filter, not just how many."""
    auth_client_a.post("/api/expenses", json=make_expense(category="Groceries", description="Kept"))
    auth_client_a.post("/api/expenses", json=make_expense(category="Dining", description="Dropped"))
    resp = auth_client_a.get("/api/export?category=Groceries")
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    assert ws.max_row == 2  # header + 1 filtered row
    assert ws.cell(2, 3).value == "Kept"


# ── Date Filtering (SG-11) ───────────────────────────────────────────


def test_export_with_date_filter(auth_client_a):
    """Export should support start_date and end_date query params."""
    today = date.today()
    yesterday = today - timedelta(days=1)
    auth_client_a.post("/api/expenses", json=make_expense(date=str(today), description="Today"))
    auth_client_a.post("/api/expenses", json=make_expense(date=str(yesterday), description="Yesterday"))
    resp = auth_client_a.get(f"/api/export?start_date={today}&end_date={today}")
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    assert ws.max_row == 2  # header + 1 row (today only)
    assert ws.cell(2, 3).value == "Today"


def test_export_with_combined_filters(auth_client_a):
    """Export with both category and date filters combined. Asserts row
    identity, not just count — three rows here fail two different ways
    (wrong category, wrong date), so a naive "any one row survived" bug
    would still report the expected count of 1."""
    today = date.today()
    yesterday = today - timedelta(days=1)
    auth_client_a.post("/api/expenses", json=make_expense(date=str(today), category="Groceries", description="Match"))
    auth_client_a.post("/api/expenses", json=make_expense(date=str(today), category="Dining", description="Wrong category"))
    auth_client_a.post("/api/expenses", json=make_expense(date=str(yesterday), category="Groceries", description="Wrong date"))
    resp = auth_client_a.get(f"/api/export?category=Groceries&start_date={today}&end_date={today}")
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    assert ws.max_row == 2  # header + 1 row (today + Groceries only)
    assert ws.cell(2, 3).value == "Match"


# ── filter_type and sort (frontend/backend export parity) ──────────────


def test_export_with_filter_type_personal(auth_client_a):
    """Export should support filter_type=personal like the list endpoint."""
    auth_client_a.post("/api/expenses", json=make_expense(split_method="Personal", description="Solo"))
    auth_client_a.post("/api/expenses", json=make_expense(split_method="50/50", description="Shared"))
    resp = auth_client_a.get("/api/export?filter_type=personal")
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    assert ws.max_row == 2  # header + 1 personal row
    assert ws.cell(2, 3).value == "Solo"


def test_export_with_filter_type_shared(auth_client_a):
    auth_client_a.post("/api/expenses", json=make_expense(split_method="Personal", description="Solo"))
    auth_client_a.post("/api/expenses", json=make_expense(split_method="50/50", description="Shared"))
    resp = auth_client_a.get("/api/export?filter_type=shared")
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    assert ws.max_row == 2  # header + 1 shared row
    assert ws.cell(2, 3).value == "Shared"


def test_export_sorted_by_amount_asc(auth_client_a):
    """Export should honor sort_by=amount&sort=asc like the list endpoint."""
    auth_client_a.post("/api/expenses", json=make_expense(amount=200))
    auth_client_a.post("/api/expenses", json=make_expense(amount=50))
    auth_client_a.post("/api/expenses", json=make_expense(amount=100))
    resp = auth_client_a.get("/api/export?sort_by=amount&sort=asc")
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    amounts = [ws.cell(r, 4).value for r in range(2, ws.max_row + 1)]
    assert amounts == [50.0, 100.0, 200.0]


def test_export_default_sort_is_date_desc(auth_client_a):
    today = date.today()
    yesterday = today - timedelta(days=1)
    auth_client_a.post("/api/expenses", json=make_expense(date=str(yesterday), description="Old"))
    auth_client_a.post("/api/expenses", json=make_expense(date=str(today), description="New"))
    resp = auth_client_a.get("/api/export")
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    descriptions = [ws.cell(r, 3).value for r in range(2, ws.max_row + 1)]
    assert descriptions == ["New", "Old"]


# ── Income sheet (07 #5 — export previously omitted Income entirely) ───


def test_export_includes_income_sheet(auth_client_a, db):
    set_mode(db, "personal")
    auth_client_a.post("/api/income", json=make_income(amount=1234, source="Salary / Wages"))
    resp = auth_client_a.get("/api/export")
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    assert "Income" in wb.sheetnames
    # The Expenses sheet must stay the active/first sheet — existing tests
    # rely on wb.active pointing at it.
    assert wb.active.title == "Expenses"

    income_ws = wb["Income"]
    headers = [cell.value for cell in income_ws[1]]
    assert headers == ["ID", "Date", "Amount", "Source", "Notes"]
    assert income_ws.cell(2, 3).value == 1234.0
    assert income_ws.cell(2, 4).value == "Salary / Wages"


def test_export_income_sheet_respects_user_isolation(auth_client_a, auth_client_b, db):
    set_mode(db, "blended")
    auth_client_a.post("/api/income", json=make_income(amount=100))
    auth_client_b.post("/api/income", json=make_income(amount=999))
    resp = auth_client_a.get("/api/export")
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    income_ws = wb["Income"]
    assert income_ws.max_row == 2  # header + Alice's own row only
    assert income_ws.cell(2, 3).value == 100.0


def test_export_income_sheet_respects_date_filter(auth_client_a, db):
    set_mode(db, "personal")
    today = str(date.today())
    last_month = str(date.today().replace(day=1) - timedelta(days=1))
    auth_client_a.post("/api/income", json=make_income(amount=100, date=today))
    auth_client_a.post("/api/income", json=make_income(amount=9999, date=last_month))
    resp = auth_client_a.get(f"/api/export?start_date={today}&end_date={today}")
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    income_ws = wb["Income"]
    assert income_ws.max_row == 2  # header + today's row only
    assert income_ws.cell(2, 3).value == 100.0
