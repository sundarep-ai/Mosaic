"""
Tests for migrate_expenses and migrate_income.

Each test spins up its own in-memory SQLite engine so migrations are fully
isolated from the shared API test database.
"""

from datetime import date
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

import bcrypt
import openpyxl
import pytest
from sqlalchemy.pool import StaticPool
from sqlmodel import Session, SQLModel, create_engine, select

import migrate_expenses
import migrate_income
from models import Expense, Income, User

# ── Shared test credentials ───────────────────────────────────────────────────

_PASSWORD_HASH = bcrypt.hashpw(b"pass", bcrypt.gensalt()).decode()
_ANSWER_HASH = bcrypt.hashpw(b"answer", bcrypt.gensalt()).decode()
_SECURITY_Q = "What is your pet's name?"


# ── Helpers ───────────────────────────────────────────────────────────────────

def _make_engine():
    """Fresh in-memory SQLite engine with all tables created."""
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    SQLModel.metadata.create_all(engine)
    return engine


def _seed_users(engine, users: list[tuple[str, str]]) -> None:
    """Seed (username, display_name) pairs into the given engine."""
    with Session(engine) as s:
        for username, display_name in users:
            s.add(User(
                username=username,
                display_name=display_name,
                password_hash=_PASSWORD_HASH,
                security_question=_SECURITY_Q,
                security_answer_hash=_ANSWER_HASH,
            ))
        s.commit()


def _write_expenses_xlsx(path: Path, rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Date", "Description", "Amount", "Category", "Paid By", "Split Method"])
    for r in rows:
        ws.append([
            r["date"], r["description"], r["amount"],
            r["category"], r["paid_by"], r["split_method"],
        ])
    wb.save(path)


def _write_income_xlsx(path: Path, rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Date", "Amount", "Source", "Display Name", "Notes"])
    for r in rows:
        ws.append([
            r["date"], r["amount"], r["source"],
            r["display_name"], r.get("notes"),
        ])
    wb.save(path)


# ── migrate_expenses ──────────────────────────────────────────────────────────

class TestMigrateExpenses:

    def _run(self, engine, xlsx_path: Path, date_format: str = "YYYY-MM-DD"):
        """Patch the module-level engine and run migrate()."""
        with (
            patch.object(migrate_expenses, "engine", engine),
            patch.object(migrate_expenses, "create_db_and_tables", lambda: None),
        ):
            migrate_expenses.migrate(str(xlsx_path), date_format)

    def test_happy_path(self, tmp_path):
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice"), ("bob", "Bob")])

        xlsx = tmp_path / "expenses.xlsx"
        _write_expenses_xlsx(xlsx, [
            {"date": "2026-01-10", "description": "Groceries", "amount": 80.00,
             "category": "Groceries", "paid_by": "Alice", "split_method": "50/50"},
            {"date": "2026-01-15", "description": "Rent", "amount": 1200.00,
             "category": "Rent", "paid_by": "Bob", "split_method": "50/50"},
        ])

        self._run(engine, xlsx)

        with Session(engine) as s:
            expenses = s.exec(select(Expense).order_by(Expense.id)).all()
        assert len(expenses) == 2
        assert expenses[0].paid_by == "Alice"
        assert expenses[0].amount == 80.00
        assert expenses[1].paid_by == "Bob"
        assert expenses[1].amount == 1200.00

    def test_date_format_iso(self, tmp_path):
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice")])
        xlsx = tmp_path / "expenses.xlsx"
        _write_expenses_xlsx(xlsx, [
            {"date": "2026-01-10", "description": "ISO", "amount": 10,
             "category": "Groceries", "paid_by": "Alice", "split_method": "Personal"},
        ])
        self._run(engine, xlsx, date_format="YYYY-MM-DD")
        with Session(engine) as s:
            expense = s.exec(select(Expense)).first()
        assert expense.date == date(2026, 1, 10)

    def test_date_format_us(self, tmp_path):
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice")])
        xlsx = tmp_path / "expenses.xlsx"
        _write_expenses_xlsx(xlsx, [
            {"date": "01/20/2026", "description": "US", "amount": 20,
             "category": "Groceries", "paid_by": "Alice", "split_method": "Personal"},
        ])
        self._run(engine, xlsx, date_format="MM/DD/YYYY")
        with Session(engine) as s:
            expense = s.exec(select(Expense)).first()
        assert expense.date == date(2026, 1, 20)

    def test_date_format_dd_mm_yyyy_does_not_swap_day_and_month(self, tmp_path):
        """Regression test for the reviewed DD/MM corruption bug: a sheet in the
        app's own default display format (DD/MM/YYYY) must import with day and
        month in the correct place for every row, including day <= 12 where the
        old guess-multiple-formats logic silently picked the wrong one.
        """
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice")])
        xlsx = tmp_path / "expenses.xlsx"
        _write_expenses_xlsx(xlsx, [
            # 3rd of November — day <= 12, so the old code would have tried
            # %m/%d/%Y first and misread this as March 11th.
            {"date": "03/11/2026", "description": "EU", "amount": 30,
             "category": "Groceries", "paid_by": "Alice", "split_method": "Personal"},
        ])
        self._run(engine, xlsx, date_format="DD/MM/YYYY")
        with Session(engine) as s:
            expense = s.exec(select(Expense)).first()
        assert expense.date == date(2026, 11, 3)

    def test_unknown_paid_by_raises(self, tmp_path):
        """Migration aborts if a paid_by value doesn't match any display name."""
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice")])

        xlsx = tmp_path / "expenses.xlsx"
        _write_expenses_xlsx(xlsx, [
            {"date": "2026-01-10", "description": "Groceries", "amount": 80.00,
             "category": "Groceries", "paid_by": "Charlie", "split_method": "50/50"},
        ])

        with pytest.raises(SystemExit) as exc_info:
            self._run(engine, xlsx)
        assert exc_info.value.code == 1

        with Session(engine) as s:
            assert s.exec(select(Expense)).first() is None

    def test_no_registered_users_raises(self, tmp_path):
        """Migration aborts if no users exist in the database."""
        engine = _make_engine()  # no users seeded

        xlsx = tmp_path / "expenses.xlsx"
        _write_expenses_xlsx(xlsx, [
            {"date": "2026-01-10", "description": "Groceries", "amount": 80.00,
             "category": "Groceries", "paid_by": "Alice", "split_method": "50/50"},
        ])

        with pytest.raises(SystemExit) as exc_info:
            self._run(engine, xlsx)
        assert exc_info.value.code == 1

    def test_missing_required_column_raises(self, tmp_path):
        """Migration aborts if a required column header is missing."""
        wb = openpyxl.Workbook()
        ws = wb.active
        # Omit "Paid By"
        ws.append(["Date", "Description", "Amount", "Category", "Split Method"])
        ws.append(["2026-01-10", "Groceries", 80, "Groceries", "50/50"])
        xlsx = tmp_path / "expenses.xlsx"
        wb.save(xlsx)

        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice")])

        with pytest.raises(SystemExit) as exc_info:
            self._run(engine, xlsx)
        assert exc_info.value.code == 1

    def test_unparseable_date_aborts_entire_import(self, tmp_path, capsys):
        """A single bad row must abort the whole import — nothing partially lands."""
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice")])

        xlsx = tmp_path / "expenses.xlsx"
        _write_expenses_xlsx(xlsx, [
            {"date": "not-a-date", "description": "Bad", "amount": 50.00,
             "category": "Groceries", "paid_by": "Alice", "split_method": "Personal"},
            {"date": "2026-01-15", "description": "Good", "amount": 90.00,
             "category": "Groceries", "paid_by": "Alice", "split_method": "Personal"},
        ])

        with pytest.raises(SystemExit) as exc_info:
            self._run(engine, xlsx)
        assert exc_info.value.code == 1

        with Session(engine) as s:
            assert s.exec(select(Expense)).all() == []

        captured = capsys.readouterr()
        assert "ERROR" in captured.out

    def test_custom_category_is_imported(self, tmp_path):
        """A novel category not in VALID_CATEGORIES must import successfully,
        matching the app's own custom-category support."""
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice")])
        xlsx = tmp_path / "expenses.xlsx"
        _write_expenses_xlsx(xlsx, [
            {"date": "2026-01-10", "description": "Good", "amount": 10,
             "category": "Groceries", "paid_by": "Alice", "split_method": "Personal"},
            {"date": "2026-01-11", "description": "Custom category", "amount": 20,
             "category": "Home Renovation", "paid_by": "Alice", "split_method": "Personal"},
        ])
        self._run(engine, xlsx)
        with Session(engine) as s:
            categories = {e.category for e in s.exec(select(Expense)).all()}
        assert categories == {"Groceries", "Home Renovation"}

    def test_empty_category_aborts_entire_import(self, tmp_path):
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice")])
        xlsx = tmp_path / "expenses.xlsx"
        _write_expenses_xlsx(xlsx, [
            {"date": "2026-01-10", "description": "Good", "amount": 10,
             "category": "Groceries", "paid_by": "Alice", "split_method": "Personal"},
            {"date": "2026-01-11", "description": "Empty category", "amount": 20,
             "category": "", "paid_by": "Alice", "split_method": "Personal"},
        ])
        with pytest.raises(SystemExit) as exc_info:
            self._run(engine, xlsx)
        assert exc_info.value.code == 1
        with Session(engine) as s:
            assert s.exec(select(Expense)).all() == []

    def test_case_collision_category_aborts_entire_import(self, tmp_path):
        """A category differing only by case from an existing one (e.g.
        'groceries' vs 'Groceries') must abort — importing it as-is would
        fragment analytics into two buckets for the same real category."""
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice")])
        xlsx = tmp_path / "expenses.xlsx"
        _write_expenses_xlsx(xlsx, [
            {"date": "2026-01-10", "description": "Good", "amount": 10,
             "category": "Groceries", "paid_by": "Alice", "split_method": "Personal"},
            {"date": "2026-01-11", "description": "Case collision", "amount": 20,
             "category": "groceries", "paid_by": "Alice", "split_method": "Personal"},
        ])
        with pytest.raises(SystemExit) as exc_info:
            self._run(engine, xlsx)
        assert exc_info.value.code == 1
        with Session(engine) as s:
            assert s.exec(select(Expense)).all() == []

    def test_unknown_split_method_aborts_entire_import(self, tmp_path):
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice"), ("bob", "Bob")])
        xlsx = tmp_path / "expenses.xlsx"
        _write_expenses_xlsx(xlsx, [
            {"date": "2026-01-10", "description": "Good", "amount": 10,
             "category": "Groceries", "paid_by": "Alice", "split_method": "50/50"},
            {"date": "2026-01-11", "description": "Bad split", "amount": 20,
             "category": "Groceries", "paid_by": "Alice", "split_method": "70/30"},
        ])
        with pytest.raises(SystemExit) as exc_info:
            self._run(engine, xlsx)
        assert exc_info.value.code == 1
        with Session(engine) as s:
            assert s.exec(select(Expense)).all() == []

    def test_amount_is_quantized_via_model_validate(self, tmp_path):
        """Amounts route through Expense.model_validate so round_amount actually runs."""
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice")])
        xlsx = tmp_path / "expenses.xlsx"
        _write_expenses_xlsx(xlsx, [
            {"date": "2026-01-10", "description": "Odd cents", "amount": 10.006,
             "category": "Groceries", "paid_by": "Alice", "split_method": "Personal"},
        ])
        self._run(engine, xlsx)
        with Session(engine) as s:
            expense = s.exec(select(Expense)).first()
        # amount is a Decimal column; comparing to a bare float (10.01 isn't
        # exactly representable in binary floating point) would spuriously
        # fail even when quantization is correct.
        assert expense.amount == Decimal("10.01")

    def test_zero_amount_aborts_entire_import(self, tmp_path):
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice")])
        xlsx = tmp_path / "expenses.xlsx"
        _write_expenses_xlsx(xlsx, [
            {"date": "2026-01-10", "description": "Good", "amount": 10,
             "category": "Groceries", "paid_by": "Alice", "split_method": "Personal"},
            {"date": "2026-01-11", "description": "Zero", "amount": 0,
             "category": "Groceries", "paid_by": "Alice", "split_method": "Personal"},
        ])
        with pytest.raises(SystemExit) as exc_info:
            self._run(engine, xlsx)
        assert exc_info.value.code == 1
        with Session(engine) as s:
            assert s.exec(select(Expense)).all() == []

    def test_existing_data_prompt_abort(self, tmp_path):
        """If the user answers N when prompted about existing data, no rows are added."""
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice")])

        # Seed one existing expense
        with Session(engine) as s:
            s.add(Expense(date=date(2025, 1, 1), description="Old", amount=10,
                          category="Groceries", paid_by="Alice", split_method="Personal"))
            s.commit()

        xlsx = tmp_path / "expenses.xlsx"
        _write_expenses_xlsx(xlsx, [
            {"date": "2026-01-10", "description": "New", "amount": 80.00,
             "category": "Groceries", "paid_by": "Alice", "split_method": "Personal"},
        ])

        with (
            patch("builtins.input", return_value="N"),
            pytest.raises(SystemExit) as exc_info,
        ):
            self._run(engine, xlsx)
        assert exc_info.value.code == 0

        with Session(engine) as s:
            expenses = s.exec(select(Expense)).all()
        assert len(expenses) == 1  # only the original
        assert expenses[0].description == "Old"

    def test_existing_data_prompt_continue(self, tmp_path):
        """If the user answers Y, new rows are appended to existing data."""
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice")])

        with Session(engine) as s:
            s.add(Expense(date=date(2025, 1, 1), description="Old", amount=10,
                          category="Groceries", paid_by="Alice", split_method="Personal"))
            s.commit()

        xlsx = tmp_path / "expenses.xlsx"
        _write_expenses_xlsx(xlsx, [
            {"date": "2026-01-10", "description": "New", "amount": 80.00,
             "category": "Groceries", "paid_by": "Alice", "split_method": "Personal"},
        ])

        with patch("builtins.input", return_value="y"):
            self._run(engine, xlsx)

        with Session(engine) as s:
            expenses = s.exec(select(Expense)).all()
        assert len(expenses) == 2

    def test_cli_requires_date_format_argument(self):
        """--date-format is mandatory at the CLI layer (argparse required=True)."""
        parser = migrate_expenses.build_arg_parser()
        with pytest.raises(SystemExit):
            parser.parse_args(["expenses.xlsx"])  # no --date-format


# ── migrate_income ────────────────────────────────────────────────────────────

class TestMigrateIncome:

    def _run(self, engine, xlsx_path: Path, date_format: str = "YYYY-MM-DD"):
        """Patch the module-level engine and run migrate()."""
        with (
            patch.object(migrate_income, "engine", engine),
            patch.object(migrate_income, "create_db_and_tables", lambda: None),
        ):
            migrate_income.migrate(str(xlsx_path), date_format)

    def test_happy_path(self, tmp_path):
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice"), ("bob", "Bob")])

        xlsx = tmp_path / "income.xlsx"
        _write_income_xlsx(xlsx, [
            {"date": "2026-01-01", "amount": 3500.00,
             "source": "Salary / Wages", "display_name": "Alice", "notes": "January salary"},
            {"date": "2026-01-05", "amount": 500.00,
             "source": "Freelance / Side Income", "display_name": "Bob", "notes": None},
        ])

        self._run(engine, xlsx)

        with Session(engine) as s:
            incomes = s.exec(select(Income)).all()
        assert len(incomes) == 2
        # Display name "Alice" must be resolved to login username "alice"
        alice_income = next(i for i in incomes if i.amount == 3500.00)
        assert alice_income.user_id == "alice"
        assert alice_income.notes == "January salary"
        bob_income = next(i for i in incomes if i.amount == 500.00)
        assert bob_income.user_id == "bob"

    def test_date_format_dd_mm_yyyy_does_not_swap_day_and_month(self, tmp_path):
        """Regression test for the reviewed DD/MM corruption bug in migrate_income."""
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice")])
        xlsx = tmp_path / "income.xlsx"
        _write_income_xlsx(xlsx, [
            {"date": "03/11/2026", "amount": 1000.00,
             "source": "Salary / Wages", "display_name": "Alice"},
        ])
        self._run(engine, xlsx, date_format="DD/MM/YYYY")
        with Session(engine) as s:
            income = s.exec(select(Income)).first()
        assert income.date == date(2026, 11, 3)

    def test_display_name_resolves_to_username(self, tmp_path):
        """user_id stored in DB must be the login username, not the display name."""
        engine = _make_engine()
        _seed_users(engine, [("praveenr", "Praveen")])

        xlsx = tmp_path / "income.xlsx"
        _write_income_xlsx(xlsx, [
            {"date": "2026-02-01", "amount": 4000.00,
             "source": "Salary / Wages", "display_name": "Praveen"},
        ])

        self._run(engine, xlsx)

        with Session(engine) as s:
            income = s.exec(select(Income)).first()
        assert income.user_id == "praveenr"
        assert income.user_id != "Praveen"

    def test_unknown_display_name_raises(self, tmp_path):
        """Migration aborts if a display name doesn't match any registered user."""
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice")])

        xlsx = tmp_path / "income.xlsx"
        _write_income_xlsx(xlsx, [
            {"date": "2026-01-01", "amount": 1000.00,
             "source": "Salary / Wages", "display_name": "Charlie"},
        ])

        with pytest.raises(SystemExit) as exc_info:
            self._run(engine, xlsx)
        assert exc_info.value.code == 1

        with Session(engine) as s:
            assert s.exec(select(Income)).first() is None

    def test_no_registered_users_raises(self, tmp_path):
        engine = _make_engine()  # no users seeded

        xlsx = tmp_path / "income.xlsx"
        _write_income_xlsx(xlsx, [
            {"date": "2026-01-01", "amount": 1000.00,
             "source": "Salary / Wages", "display_name": "Alice"},
        ])

        with pytest.raises(SystemExit) as exc_info:
            self._run(engine, xlsx)
        assert exc_info.value.code == 1

    def test_missing_required_column_raises(self, tmp_path):
        """Migration aborts if a required column header is missing."""
        wb = openpyxl.Workbook()
        ws = wb.active
        # Omit "Display Name"
        ws.append(["Date", "Amount", "Source", "Notes"])
        ws.append(["2026-01-01", 1000, "Salary / Wages", None])
        xlsx = tmp_path / "income.xlsx"
        wb.save(xlsx)

        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice")])

        with pytest.raises(SystemExit) as exc_info:
            self._run(engine, xlsx)
        assert exc_info.value.code == 1

    def test_invalid_source_aborts_entire_import(self, tmp_path, capsys):
        """A row with an unrecognised income source aborts the whole import."""
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice")])

        xlsx = tmp_path / "income.xlsx"
        _write_income_xlsx(xlsx, [
            {"date": "2026-01-01", "amount": 500.00,
             "source": "Gambling Winnings", "display_name": "Alice"},
            {"date": "2026-01-02", "amount": 3000.00,
             "source": "Salary / Wages", "display_name": "Alice"},
        ])

        with pytest.raises(SystemExit) as exc_info:
            self._run(engine, xlsx)
        assert exc_info.value.code == 1

        with Session(engine) as s:
            assert s.exec(select(Income)).all() == []

        captured = capsys.readouterr()
        assert "ERROR" in captured.out

    def test_non_positive_amount_aborts_entire_import(self, tmp_path):
        """Rows with zero or negative amounts abort the whole import."""
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice")])

        xlsx = tmp_path / "income.xlsx"
        _write_income_xlsx(xlsx, [
            {"date": "2026-01-01", "amount": 0,
             "source": "Salary / Wages", "display_name": "Alice"},
            {"date": "2026-01-02", "amount": -100,
             "source": "Salary / Wages", "display_name": "Alice"},
            {"date": "2026-01-03", "amount": 2000,
             "source": "Salary / Wages", "display_name": "Alice"},
        ])

        with pytest.raises(SystemExit) as exc_info:
            self._run(engine, xlsx)
        assert exc_info.value.code == 1

        with Session(engine) as s:
            assert s.exec(select(Income)).all() == []

    def test_empty_display_name_aborts_entire_import(self, tmp_path):
        """A row with an empty display name cell aborts the whole import."""
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice")])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Date", "Amount", "Source", "Display Name", "Notes"])
        ws.append(["2026-01-01", 1000, "Salary / Wages", "", None])  # empty display name
        ws.append(["2026-01-02", 2000, "Salary / Wages", "Alice", None])
        xlsx = tmp_path / "income.xlsx"
        wb.save(xlsx)

        with pytest.raises(SystemExit) as exc_info:
            self._run(engine, xlsx)
        assert exc_info.value.code == 1

        with Session(engine) as s:
            assert s.exec(select(Income)).all() == []

    def test_unparseable_date_aborts_entire_import(self, tmp_path, capsys):
        """Rows with unparseable dates abort the whole import."""
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice")])

        xlsx = tmp_path / "income.xlsx"
        _write_income_xlsx(xlsx, [
            {"date": "not-a-date", "amount": 500.00,
             "source": "Salary / Wages", "display_name": "Alice"},
            {"date": "2026-01-15", "amount": 3000.00,
             "source": "Salary / Wages", "display_name": "Alice"},
        ])

        with pytest.raises(SystemExit) as exc_info:
            self._run(engine, xlsx)
        assert exc_info.value.code == 1

        with Session(engine) as s:
            assert s.exec(select(Income)).all() == []

        captured = capsys.readouterr()
        assert "ERROR" in captured.out

    def test_existing_data_prompt_abort(self, tmp_path):
        """If the user answers N when prompted about existing data, no rows are added."""
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice")])

        with Session(engine) as s:
            s.add(Income(date=date(2025, 1, 1), amount=1000, source="Salary / Wages",
                         notes=None, user_id="alice"))
            s.commit()

        xlsx = tmp_path / "income.xlsx"
        _write_income_xlsx(xlsx, [
            {"date": "2026-01-01", "amount": 2000.00,
             "source": "Salary / Wages", "display_name": "Alice"},
        ])

        with (
            patch("builtins.input", return_value="N"),
            pytest.raises(SystemExit) as exc_info,
        ):
            self._run(engine, xlsx)
        assert exc_info.value.code == 0

        with Session(engine) as s:
            incomes = s.exec(select(Income)).all()
        assert len(incomes) == 1
        assert incomes[0].amount == 1000

    def test_existing_data_prompt_continue(self, tmp_path):
        """If the user answers Y, new rows are appended to existing data."""
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice")])

        with Session(engine) as s:
            s.add(Income(date=date(2025, 1, 1), amount=1000, source="Salary / Wages",
                         notes=None, user_id="alice"))
            s.commit()

        xlsx = tmp_path / "income.xlsx"
        _write_income_xlsx(xlsx, [
            {"date": "2026-01-01", "amount": 2000.00,
             "source": "Salary / Wages", "display_name": "Alice"},
        ])

        with patch("builtins.input", return_value="y"):
            self._run(engine, xlsx)

        with Session(engine) as s:
            incomes = s.exec(select(Income)).all()
        assert len(incomes) == 2

    def test_old_user_id_column_header_still_works(self, tmp_path):
        """Sheets with a 'User ID' column header (old format) are still accepted."""
        engine = _make_engine()
        _seed_users(engine, [("alice", "Alice")])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Date", "Amount", "Source", "User ID", "Notes"])
        ws.append(["2026-01-01", 3000, "Salary / Wages", "Alice", None])
        xlsx = tmp_path / "income.xlsx"
        wb.save(xlsx)

        self._run(engine, xlsx)

        with Session(engine) as s:
            income = s.exec(select(Income)).first()
        assert income is not None
        assert income.user_id == "alice"

    def test_cli_requires_date_format_argument(self):
        """--date-format is mandatory at the CLI layer (argparse required=True)."""
        parser = migrate_income.build_arg_parser()
        with pytest.raises(SystemExit):
            parser.parse_args(["income.xlsx"])  # no --date-format
