from datetime import date
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, Request
from fastapi.responses import StreamingResponse
from sqlalchemy import or_
from sqlmodel import Session, select
import openpyxl

from auth import get_current_user
from database import get_session
from models import Expense, Income
from utils import escape_like as _escape_like

router = APIRouter()


@router.get("/export")
def export_expenses(
    request: Request,
    search: Optional[str] = None,
    paid_by: Optional[str] = None,
    category: Optional[str] = None,
    filter_type: Optional[str] = None,
    sort: Optional[str] = "desc",
    sort_by: Optional[str] = "date",
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    session: Session = Depends(get_session),
    current_user: str = Depends(get_current_user),
):
    statement = select(Expense)

    if search:
        escaped = _escape_like(search)
        statement = statement.where(
            or_(
                Expense.description.like(f"%{escaped}%", escape="\\"),
                Expense.category.like(f"%{escaped}%", escape="\\"),
                Expense.paid_by.like(f"%{escaped}%", escape="\\"),
            )
        )
    if paid_by:
        statement = statement.where(Expense.paid_by == paid_by)
    if category:
        statement = statement.where(Expense.category == category)
    if filter_type == "personal":
        statement = statement.where(Expense.split_method == "Personal")
    elif filter_type == "shared":
        statement = statement.where(Expense.split_method != "Personal")
    if start_date:
        statement = statement.where(Expense.date >= start_date)
    if end_date:
        statement = statement.where(Expense.date <= end_date)

    order_col = Expense.amount if sort_by == "amount" else Expense.date
    statement = statement.order_by(order_col.desc()) if sort == "desc" else statement.order_by(order_col.asc())

    expenses = session.exec(statement).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"

    headers = ["ID", "Date", "Description", "Amount", "Category", "Paid By", "Split Method"]
    ws.append(headers)

    # Style header row
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for e in expenses:
        ws.append([e.id, str(e.date), e.description, e.amount, e.category, e.paid_by, e.split_method])

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Income sheet — a "full snapshot" export must not silently omit income.
    # search/paid_by/category/filter_type/sort are expense-specific concepts
    # that don't map onto Income; only the shared date range applies to both.
    income_statement = select(Income).where(Income.user_id == current_user)
    if start_date:
        income_statement = income_statement.where(Income.date >= start_date)
    if end_date:
        income_statement = income_statement.where(Income.date <= end_date)
    income_statement = income_statement.order_by(Income.date.desc())
    income_entries = session.exec(income_statement).all()

    income_ws = wb.create_sheet("Income")
    income_ws.append(["ID", "Date", "Amount", "Source", "Notes"])
    for cell in income_ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for inc in income_entries:
        income_ws.append([inc.id, str(inc.date), inc.amount, inc.source, inc.notes or ""])
    for col in income_ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        income_ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.spreadsheet",
        headers={"Content-Disposition": f"attachment; filename=expenses_{date.today().isoformat()}.xlsx"},
    )
